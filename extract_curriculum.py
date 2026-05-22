import openpyxl
import json
import re
import os
from collections import defaultdict

def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()

def extract_curriculum_mapping(filepath):
    """Extract all data from the Curriculum Mapping file.
    Note: Column A is empty in this file; data starts at column B (index 1)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {
        "overview": {},
        "domains": {},
        "dok_framework": {},
        "academic_calendar": [],
        "domain_priority": [],
        "domain_coverage": [],
        "scope_and_sequence_consolidated": [],
        "assessment_calendar": [],
        "grades": {}
    }

    # === OVERVIEW SHEET ===
    ws = wb['Overview']
    
    # Extract domains S1-S9 (data in cols B, C, D)
    for row in ws.iter_rows(min_row=20, max_row=28, values_only=True):
        code = safe_str(row[1])
        name = safe_str(row[2])
        desc = safe_str(row[3])
        if code.startswith('S') and name:
            data["domains"][code] = {
                "name": name,
                "description": desc
            }

    # DOK Framework (rows 38-41, data in cols B, C, D, E)
    for row in ws.iter_rows(min_row=38, max_row=41, values_only=True):
        level = safe_str(row[1])
        demand = safe_str(row[2])
        desc = safe_str(row[3])
        target = safe_str(row[4])
        if level.startswith('DOK'):
            data["dok_framework"][level] = {
                "cognitive_demand": demand,
                "description": desc,
                "target_percentage": target
            }

    # Academic Calendar (rows 13-16, data in cols B-J)
    for row in ws.iter_rows(min_row=13, max_row=16, values_only=True):
        term_name = safe_str(row[1])
        if term_name.startswith('Term') or term_name == 'TOTAL':
            data["academic_calendar"].append({
                "term": term_name,
                "start_date": safe_str(row[2]),
                "end_date": safe_str(row[3]),
                "total_weeks": safe_str(row[4]),
                "instructional_weeks": safe_str(row[5]),
                "mid_term_exams": safe_str(row[6]),
                "mid_term_break": safe_str(row[7]),
                "final_exams": safe_str(row[8]),
                "lessons_45min": safe_str(row[9])
            })

    # Domain Priority Framework (rows 32-34, data in cols B-D)
    for row in ws.iter_rows(min_row=32, max_row=34, values_only=True):
        term = safe_str(row[1])
        c1 = safe_str(row[2])
        c2 = safe_str(row[3])
        g9 = safe_str(row[4])
        if term.startswith('Term'):
            data["domain_priority"].append({
                "term": term,
                "cycle_1_g1_g5": c1,
                "cycle_2_g6_g8": c2,
                "grade_9": g9
            })

    # === GRADE SHEETS ===
    grade_sheets = [f'Grade {i}' for i in range(1, 10)]
    for sheet_name in grade_sheets:
        ws = wb[sheet_name]
        grade_num = sheet_name.replace('Grade ', '')
        grade_key = f"G{grade_num}"
        
        lessons = []
        summary = ""
        teacher_info = {}
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            val_b = safe_str(row[1]) if len(row) > 1 else ""
            
            if i <= 5:
                if "Cycle" in val_b:
                    teacher_info["cycle_info"] = val_b
                elif "Mr." in val_b:
                    teacher_info["teacher"] = val_b
                continue
            if i == 6:
                continue
            
            week = row[1] if len(row) > 1 else None
            
            if week is None:
                if val_b and "Summary" in val_b:
                    summary = val_b
                continue
            
            lessons.append({
                "week_number": week,
                "date_range": safe_str(row[2]) if len(row) > 2 else "",
                "term": safe_str(row[3]) if len(row) > 3 else "",
                "phase": safe_str(row[4]) if len(row) > 4 else "",
                "unit": safe_str(row[5]) if len(row) > 5 else "",
                "lesson_title": safe_str(row[6]) if len(row) > 6 else "",
                "domains": safe_str(row[7]) if len(row) > 7 else "",
                "slo_codes": safe_str(row[8]) if len(row) > 8 else "",
                "dok_level": safe_str(row[9]) if len(row) > 9 else "",
                "learning_activity": safe_str(row[10]) if len(row) > 10 else "",
                "assessment": safe_str(row[11]) if len(row) > 11 else "",
                "resources": safe_str(row[12]) if len(row) > 12 else ""
            })
        
        data["grades"][grade_key] = {
            "grade_number": int(grade_num),
            "teacher_info": teacher_info,
            "summary": summary,
            "lessons": lessons
        }

    # === DOMAIN COVERAGE SHEET ===
    ws = wb['Domain Coverage']
    domain_keys = ['S1_History', 'S2_Civics', 'S3_Geography', 'S4_Sociology', 
                   'S5_Economics', 'S6_Info_Lit', 'S7_Info_Proc', 'S8_Moral_Ed', 'S9_UAE_Culture']
    for i, row in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True), 6):
        grade = safe_str(row[1])
        cycle = safe_str(row[2])
        term = safe_str(row[3])
        if not grade or not grade.startswith('G'):
            continue
        coverage = {}
        for j, dk in enumerate(domain_keys):
            val = safe_str(row[4+j]) if (4+j) < len(row) and row[4+j] is not None else ""
            coverage[dk] = (val == '✓')
        data["domain_coverage"].append({
            "grade": grade,
            "cycle": cycle,
            "term": term,
            "coverage": coverage
        })

    # === SCOPE & SEQUENCE CONSOLIDATED SHEET ===
    ws = wb['Scope & Sequence']
    for i, row in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True), 6):
        wk = row[1] if len(row) > 1 else None
        if wk is None:
            continue
        data["scope_and_sequence_consolidated"].append({
            "week": wk,
            "date_range": safe_str(row[2]) if len(row) > 2 else "",
            "term": safe_str(row[3]) if len(row) > 3 else "",
            "phase": safe_str(row[4]) if len(row) > 4 else "",
            "g1_g3": safe_str(row[5]) if len(row) > 5 else "",
            "g4_g5": safe_str(row[6]) if len(row) > 6 else "",
            "g6": safe_str(row[7]) if len(row) > 7 else "",
            "g7_g8": safe_str(row[8]) if len(row) > 8 else "",
            "g9": safe_str(row[9]) if len(row) > 9 else "",
            "assessment": safe_str(row[10]) if len(row) > 10 else "",
            "key_events": safe_str(row[11]) if len(row) > 11 else ""
        })

    # === ASSESSMENT CALENDAR SHEET ===
    ws = wb['Assessment Calendar']
    for i, row in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True), 6):
        num = row[1] if len(row) > 1 else None
        if num is None:
            continue
        # Only include rows that are actual assessments (numbered)
        if not isinstance(num, (int, float)):
            continue
        data["assessment_calendar"].append({
            "number": num,
            "assessment": safe_str(row[2]) if len(row) > 2 else "",
            "term": safe_str(row[3]) if len(row) > 3 else "",
            "type": safe_str(row[4]) if len(row) > 4 else "",
            "dok_range": safe_str(row[5]) if len(row) > 5 else "",
            "date_window": safe_str(row[6]) if len(row) > 6 else "",
            "g1_g3": safe_str(row[7]) if len(row) > 7 else "",
            "g4_g5": safe_str(row[8]) if len(row) > 8 else "",
            "g6": safe_str(row[9]) if len(row) > 9 else "",
            "g7_g8": safe_str(row[10]) if len(row) > 10 else "",
            "g9": safe_str(row[11]) if len(row) > 11 else ""
        })

    wb.close()
    return data


def extract_scope_sequence(filepath, term_num):
    """Extract data from a Scope & Sequence Term file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Determine grade from sheet name
        grade_match = re.search(r'Grade\s+(\d+)', sheet_name)
        if not grade_match:
            # Handle "Standards Reference" or other non-grade sheets
            if 'Standard' in sheet_name or 'Reference' in sheet_name:
                stds = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                    vals = [safe_str(v) for v in row if v is not None]
                    if vals:
                        stds.append(vals)
                data["Standards_Reference"] = stds
                continue
            continue
        
        grade_num = grade_match.group(1)
        grade_key = f"G{grade_num}"
        
        # Determine header row by scanning for the row with 'Week' header
        header_row = None
        col_map = {}
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_strs = [safe_str(v).lower() for v in row]
            combined = ' '.join(row_strs)
            if 'week' in combined and ('domain' in combined or 'theme' in combined or 'lesson' in combined):
                header_row = i
                for j, v in enumerate(row):
                    sv = safe_str(v).lower()
                    if 'week' in sv:
                        col_map['week'] = j
                    elif 'date' in sv:
                        col_map['date'] = j
                    elif 'domain' in sv and 'learning' in sv:
                        col_map['domain'] = j
                    elif 'theme' in sv or 'lesson title' in sv:
                        col_map['lesson_title'] = j
                    elif 'standard' in sv:
                        col_map['standard'] = j
                    elif 'objective' in sv or 'swbat' in sv:
                        col_map['objective'] = j
                    elif 'success' in sv and 'criteria' in sv:
                        col_map['success_criteria'] = j
                    elif 'prior' in sv or 'engagement' in sv:
                        col_map['prior_learning'] = j
                    elif 'teaching' in sv or 'learning activit' in sv:
                        col_map['activities'] = j
                    elif 'assessment' in sv or 'closure' in sv:
                        col_map['assessment'] = j
                    elif 'resource' in sv or 'material' in sv:
                        col_map['resources'] = j
                    elif 'homework' in sv or 'extension' in sv:
                        col_map['homework'] = j
                break
        
        if header_row is None:
            header_row = 6 if term_num <= 2 else 7
            # Default col map - scan manually
            # For T3, the columns are slightly different
            for j, v in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__()):
                sv = safe_str(v).lower()
                if 'week' in sv:
                    col_map['week'] = j
                elif 'date' in sv:
                    col_map['date'] = j
                elif 'domain' in sv:
                    col_map['domain'] = j
                elif 'theme' in sv or 'lesson' in sv:
                    col_map['lesson_title'] = j
                elif 'standard' in sv:
                    col_map['standard'] = j
                elif 'objective' in sv or 'swbat' in sv:
                    col_map['objective'] = j
                elif 'success' in sv:
                    col_map['success_criteria'] = j
                elif 'prior' in sv:
                    col_map['prior_learning'] = j
                elif 'teaching' in sv or 'activit' in sv:
                    col_map['activities'] = j
                elif 'assessment' in sv:
                    col_map['assessment'] = j
                elif 'resource' in sv or 'material' in sv:
                    col_map['resources'] = j
                elif 'homework' in sv or 'extension' in sv:
                    col_map['homework'] = j
        
        lessons = []
        header_info = {}
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            if i < header_row:
                val = safe_str(row[1]) if len(row) > 1 else safe_str(row[0]) if len(row) > 0 else ""
                if "MSCS" in val and "Scope" in val:
                    header_info["title"] = val
                elif "Term" in val and ("Cycle" in val or "1 x" in val or "1 Class" in val):
                    header_info["term_info"] = val
                elif "Mr." in val:
                    header_info["teacher"] = val
                elif "Textbook" in val:
                    header_info["textbook"] = val
                elif "Ain Al Khaleej" in val:
                    header_info["school"] = val
                continue
            elif i == header_row:
                continue
            
            def get_col(name):
                idx = col_map.get(name)
                if idx is not None and idx < len(row):
                    return safe_str(row[idx])
                return ""
            
            week_val = get_col('week')
            if not week_val:
                continue
            
            # Skip summary/footer rows
            if 'Summary' in str(week_val) or 'Prepared by' in str(week_val) or 'Academic Year' in str(week_val):
                continue
            
            lesson = {
                "week": week_val,
                "date": get_col('date'),
                "learning_domain": get_col('domain'),
                "lesson_title": get_col('lesson_title'),
                "moe_standard": get_col('standard'),
                "learning_objective": get_col('objective'),
                "success_criteria": get_col('success_criteria'),
                "prior_learning_engagement": get_col('prior_learning'),
                "teaching_learning_activities": get_col('activities'),
                "assessment_closure": get_col('assessment'),
                "resources_materials": get_col('resources'),
                "homework_extension": get_col('homework')
            }
            
            lessons.append(lesson)
        
        data[grade_key] = {
            "grade_number": int(grade_num),
            "term": term_num,
            "header_info": header_info,
            "lessons": lessons
        }
    
    wb.close()
    return data


def match_ss_to_mapping(mapping_lesson, ss_lessons, term_key, grade_key):
    """Match a scope & sequence lesson to a curriculum mapping lesson.
    
    Week number mapping:
    - T1: S&S week N = Mapping week N (same)
    - T2: S&S week N = Mapping week (N + 14)
    - T3: S&S week N = Mapping week (N + 26)
    """
    term_offsets = {"T1": 0, "T2": 14, "T3": 26}
    offset = term_offsets.get(term_key, 0)
    
    map_week = mapping_lesson["week_number"]
    try:
        map_week_num = int(map_week)
    except (ValueError, TypeError):
        return None
    
    expected_ss_week = map_week_num - offset
    
    for ss_lesson in ss_lessons:
        ss_week = ss_lesson.get("week", "")
        try:
            # Extract numeric week from "Week 1" or "1" etc.
            ss_week_match = re.search(r'(\d+)', str(ss_week))
            ss_week_num = int(ss_week_match.group(1)) if ss_week_match else None
        except:
            ss_week_num = None
        
        if ss_week_num is not None and ss_week_num == expected_ss_week:
            return ss_lesson
    
    return None


def build_structured_output(mapping_data, ss_data):
    """Build the final structured JSON: grade -> term -> unit -> lesson with ALL fields."""
    structured = {
        "metadata": {
            "academic_year": "2026-2027",
            "subject": "Moral, Social, and Cultural Studies (MSCS)",
            "curriculum": "English-Medium American Curriculum",
            "jurisdiction": "ADEK",
            "teacher": "Mr. Ahmed Mahmoud Saeed Ahmed Ali",
            "email": "ahmed.ali.academic@gmail.com",
            "location": "Al Ain, UAE",
            "school": "Ain Al Khaleej Private School, Al Ain, UAE",
            "scheduling": "45 minutes per week (1 lesson)",
            "total_instructional_lessons_per_year": 24,
            "term_1_lessons": 10,
            "term_2_lessons": 7,
            "term_3_lessons": 7,
            "total_annual_instruction_hours": 18,
            "critical_constraint": "Only 45 minutes per week (1 lesson) — Every lesson must be strategically planned for maximum impact",
            "domains": mapping_data["domains"],
            "dok_framework": mapping_data["dok_framework"],
            "academic_calendar": mapping_data["academic_calendar"],
            "domain_priority_framework": mapping_data["domain_priority"]
        },
        "domain_coverage_matrix": mapping_data["domain_coverage"],
        "consolidated_scope_and_sequence": mapping_data["scope_and_sequence_consolidated"],
        "assessment_calendar": mapping_data["assessment_calendar"],
        "grades": {}
    }
    
    # Build grade -> term -> unit -> lesson structure
    for grade_key in sorted(mapping_data["grades"].keys(), key=lambda x: int(x[1:])):
        grade_data = mapping_data["grades"][grade_key]
        grade_num = grade_data["grade_number"]
        
        if grade_num <= 5:
            cycle = "Cycle 1 (G1-G5)"
        elif grade_num <= 8:
            cycle = "Cycle 2 (G6-G8)"
        else:
            cycle = "Cycle 2 (Grade 9)"
        
        grade_entry = {
            "grade_number": grade_num,
            "grade_label": f"Grade {grade_num}",
            "cycle": cycle,
            "summary": grade_data["summary"],
            "teacher_info": grade_data["teacher_info"],
            "terms": {}
        }
        
        # Group lessons by term then by unit
        term_groups = defaultdict(lambda: defaultdict(list))
        
        for lesson in grade_data["lessons"]:
            term = lesson["term"]
            unit = lesson["unit"] if lesson["unit"] and lesson["unit"] not in ["—", ""] else "General"
            term_groups[term][unit].append(lesson)
        
        # Get S&S data for this grade
        ss_grade_data = {}
        for term_num in [1, 2, 3]:
            if term_num in ss_data and grade_key in ss_data[term_num]:
                ss_grade_data[term_num] = ss_data[term_num][grade_key]
        
        # Build term -> unit -> lesson structure
        for term_key in ["T1", "T2", "T3"]:
            if term_key not in term_groups:
                continue
            
            term_num = {"T1": 1, "T2": 2, "T3": 3}.get(term_key)
            ss_lessons = []
            if term_num in ss_grade_data:
                ss_lessons = ss_grade_data[term_num].get("lessons", [])
            
            units_dict = {}
            for unit_key, unit_lessons in term_groups[term_key].items():
                unit_lesson_list = []
                for lesson in unit_lessons:
                    lesson_entry = {
                        "week_number": lesson["week_number"],
                        "date_range": lesson["date_range"],
                        "phase": lesson["phase"],
                        "lesson_title": lesson["lesson_title"],
                        "domains": lesson["domains"],
                        "slo_codes": lesson["slo_codes"],
                        "dok_level": lesson["dok_level"],
                        "learning_activity": lesson["learning_activity"],
                        "assessment": lesson["assessment"],
                        "resources": lesson["resources"]
                    }
                    
                    # Match with Scope & Sequence detail
                    ss_match = match_ss_to_mapping(lesson, ss_lessons, term_key, grade_key)
                    if ss_match:
                        lesson_entry["scope_sequence_detail"] = {
                            "learning_objective": ss_match.get("learning_objective", ""),
                            "success_criteria": ss_match.get("success_criteria", ""),
                            "prior_learning_engagement": ss_match.get("prior_learning_engagement", ""),
                            "teaching_learning_activities": ss_match.get("teaching_learning_activities", ""),
                            "assessment_closure": ss_match.get("assessment_closure", ""),
                            "resources_materials": ss_match.get("resources_materials", ""),
                            "homework_extension": ss_match.get("homework_extension", ""),
                            "moe_standard": ss_match.get("moe_standard", ""),
                            "learning_domain": ss_match.get("learning_domain", "")
                        }
                    
                    unit_lesson_list.append(lesson_entry)
                
                units_dict[unit_key] = {
                    "unit_name": unit_key,
                    "lesson_count": len(unit_lesson_list),
                    "lessons": unit_lesson_list
                }
            
            grade_entry["terms"][term_key] = {
                "term_name": f"Term {term_key[1]}",
                "units": units_dict
            }
        
        structured["grades"][grade_key] = grade_entry
    
    # Also add full scope & sequence data separately for complete reference
    structured["scope_sequence_details"] = {}
    for ss_term_num, ss_content in ss_data.items():
        term_key = f"Term_{ss_term_num}"
        structured["scope_sequence_details"][term_key] = ss_content
    
    return structured


# === MAIN EXECUTION ===
print("Step 1: Extracting Curriculum Mapping...")
mapping_data = extract_curriculum_mapping('upload/MSCS_Curriculum_Mapping_2026-2027.xlsx')
print(f"  Grades: {len(mapping_data['grades'])}")
for gk in sorted(mapping_data['grades'].keys(), key=lambda x: int(x[1:])):
    print(f"    {gk}: {len(mapping_data['grades'][gk]['lessons'])} lessons")
print(f"  Domains: {len(mapping_data['domains'])} -> {list(mapping_data['domains'].keys())}")
print(f"  DOK Framework: {len(mapping_data['dok_framework'])} levels")
print(f"  Academic Calendar: {len(mapping_data['academic_calendar'])} entries")
print(f"  Domain Coverage: {len(mapping_data['domain_coverage'])} entries")
print(f"  Assessment Calendar: {len(mapping_data['assessment_calendar'])} entries")
print(f"  Consolidated S&S: {len(mapping_data['scope_and_sequence_consolidated'])} entries")

print("\nStep 2: Extracting Scope & Sequence Term 1...")
ss_t1 = extract_scope_sequence('upload/MSCS_Scope_Sequence_Term1_2026-2027.xlsx', 1)
for gk in sorted(ss_t1.keys()):
    if isinstance(ss_t1[gk], dict) and 'lessons' in ss_t1[gk]:
        print(f"  {gk}: {len(ss_t1[gk]['lessons'])} lessons")

print("\nStep 3: Extracting Scope & Sequence Term 2...")
ss_t2 = extract_scope_sequence('upload/MSCS_Scope_Sequence_Term2_2026-2027.xlsx', 2)
for gk in sorted(ss_t2.keys()):
    if isinstance(ss_t2[gk], dict) and 'lessons' in ss_t2[gk]:
        print(f"  {gk}: {len(ss_t2[gk]['lessons'])} lessons")

print("\nStep 4: Extracting Scope & Sequence Term 3...")
ss_t3 = extract_scope_sequence('upload/MSCS_Scope_Sequence_Term3_2026-2027.xlsx', 3)
for gk in sorted(ss_t3.keys()):
    if isinstance(ss_t3[gk], dict) and 'lessons' in ss_t3[gk]:
        print(f"  {gk}: {len(ss_t3[gk]['lessons'])} lessons")

# Combine S&S data
ss_data = {1: ss_t1, 2: ss_t2, 3: ss_t3}

print("\nStep 5: Building structured output...")
structured = build_structured_output(mapping_data, ss_data)

# Save to JSON
output_path = 'src/lib/curriculum_mapping.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(structured, f, indent=2, ensure_ascii=False, default=str)

file_size = os.path.getsize(output_path)
print(f"\nSaved to {output_path}")
print(f"File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")

# Print comprehensive summary
print("\n" + "="*80)
print("COMPLETE EXTRACTION SUMMARY")
print("="*80)

print(f"\n📊 OVERVIEW")
print(f"  Academic Year: 2026-2027")
print(f"  Subject: MSCS - Moral, Social, and Cultural Studies")
print(f"  Curriculum: English-Medium American, ADEK Jurisdiction")
print(f"  Schedule: 45 min/week, 24 instructional lessons/year")

print(f"\n📚 DOMAINS (S1-S9)")
for code, info in structured["metadata"]["domains"].items():
    print(f"  {code}: {info['name']} - {info['description'][:60]}...")

print(f"\n📈 DOK FRAMEWORK")
for level, info in structured["metadata"]["dok_framework"].items():
    print(f"  {level}: {info['cognitive_demand']} ({info['target_percentage']})")

print(f"\n📅 ACADEMIC CALENDAR")
for cal in structured["metadata"]["academic_calendar"]:
    print(f"  {cal['term']}: {cal.get('start_date','')} to {cal.get('end_date','')} ({cal.get('lessons_45min','')} lessons)")

total_lessons = 0
total_with_ss = 0
print(f"\n🎓 GRADES (Curriculum Mapping + Scope & Sequence)")
for gk in sorted(structured["grades"].keys(), key=lambda x: int(x[1:])):
    gv = structured["grades"][gk]
    grade_lessons = 0
    grade_ss = 0
    terms_info = []
    for tk, tv in gv["terms"].items():
        term_lessons = sum(len(uv["lessons"]) for uv in tv["units"].values())
        ss_count = sum(1 for uv in tv["units"].values() for l in uv["lessons"] if "scope_sequence_detail" in l)
        grade_lessons += term_lessons
        grade_ss += ss_count
        terms_info.append(f"{tk}:{term_lessons}L/{ss_count}SS")
    total_lessons += grade_lessons
    total_with_ss += grade_ss
    print(f"  {gk} (Grade {gv['grade_number']}, {gv['cycle']}): {grade_lessons} total lessons, {grade_ss} with S&S detail [{', '.join(terms_info)}]")

print(f"\n  TOTAL: {total_lessons} curriculum mapping lessons, {total_with_ss} enriched with S&S detail")

print(f"\n📋 ASSESSMENT CALENDAR ({len(structured['assessment_calendar'])} entries)")
for ac in structured['assessment_calendar']:
    print(f"  #{ac['number']}: {ac['assessment']} ({ac['term']}, {ac['type']}, {ac['dok_range']})")

print(f"\n🌐 DOMAIN COVERAGE MATRIX ({len(structured['domain_coverage_matrix'])} entries)")
print(f"  Covering {len(set(dc['grade'] for dc in structured['domain_coverage_matrix']))} grades x 3 terms")

print(f"\n📖 SCOPE & SEQUENCE DETAILS")
for tk, tv in structured["scope_sequence_details"].items():
    grade_count = sum(1 for k in tv.keys() if k.startswith('G'))
    total_ss_lessons = sum(len(tv[k]['lessons']) for k in tv.keys() if isinstance(tv[k], dict) and 'lessons' in tv[k])
    print(f"  {tk}: {grade_count} grades, {total_ss_lessons} total lessons with full detail")
